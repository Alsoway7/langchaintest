# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Status"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="4472C4")
GRAY   = PatternFill("solid", fgColor="D9D9D9")
bold_white = Font(bold=True, color="FFFFFF")
bold_black = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def c(ws, row, col, val, fill=None, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=val)
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    cell.border = border
    return cell

# ===== ヘッダー =====
for col, h in enumerate(["質問例", "質問内容", "回答状態", "理由・必要データ"], 1):
    c(ws, 1, col, h, fill=BLUE, font=bold_white, align=center)

rows = [
    ("例1\nExcel参照",
     "「4_相同性が高い生物種リスト(BLAST結果10位まで)」Excel参照\nr58-240918S のリード数合計を記載",
     "❌ 回答不可",
     "Excel に r58-240918S 列が存在しない\n現在の列: dog-food-rbcL のみ（gPlant）\n必要: 論文用 gPlant Excel（r58等の列含む）"),

    ("例1\nExcel参照",
     "r58-240918S のリード数≠0 の ASV 番号をすべて列挙",
     "❌ 回答不可",
     "同上"),

    ("例1\nExcel参照",
     "リード数≠0 の ASV に帰属される植物種（学名）を記載",
     "❌ 回答不可",
     "同上"),

    ("例1\nExcel参照",
     "リード数≠0 の ASV 表 → 同一植物種ごとにまとめた表を作成",
     "❌ 回答不可",
     "同上"),

    ("例2\n論文参照",
     "修士論文参照：r3-230606S の採取地名・緯度経度・\n試料区分・ミツバチ種を回答",
     "✅ 回答可能",
     "論文 DOCX から取得済み\n採取地: 高崎市上豊岡町\n座標: 36.331, 138.973\n試料: 巣くず(Nest) / セイヨウミツバチ(mellifera)"),

    ("例2\n論文参照",
     "修士論文参照：r26-221010M の採取地名・緯度経度・\n試料区分・ミツバチ種を回答",
     "✅ 回答可能",
     "論文 DOCX から取得済み\n採取地: 高崎市上豊岡町\n座標: 36.331, 138.973\n試料: 蜂蜜(Honey) / ニホンミツバチ(japonica)"),

    ("例3\nFASTA参照",
     "r3-230606S に含まれる FASTA ファイル数を回答",
     "❌ 回答不可",
     "230606S.fasta が data/ に存在しない\n必要: data/04_sequences_fasta__gPlant__230606S.fasta"),

    ("例4\nBLAST参照",
     "r3-230606S の各 ASV → NCBI BLAST → 植物種上位10位表\n→ 上位1位表 → 植物種ごとまとめ表（3つの表を作成）",
     "❌ 回答不可",
     "FASTA ファイル不在のため BLAST 不可\n※FASTA が揃えば NCBI BLAST 自動呼出し機能は実装済み"),

    ("例3\nFASTA参照",
     "r26-221010M に含まれる FASTA ファイル数を回答",
     "❌ 回答不可",
     "221010M.fasta が data/ に存在しない\n必要: data/04_sequences_fasta__gPlant__221010M.fasta"),

    ("例4\nBLAST参照",
     "r26-221010M の各 ASV → NCBI BLAST → 植物種上位10位表\n→ 上位1位表 → 植物種ごとまとめ表（3つの表を作成）",
     "❌ 回答不可",
     "同上"),

    ("例5\n比較",
     "r58-240918S と r26-221010M の比較\n共通植物種・片方のみの植物種を回答\n(a) Excel参照 または (b) 例4の表をもとに",
     "❌ 回答不可",
     "どちらの試料も Excel・FASTA ともに未提供\n(a): 論文用 Excel（r58・r26 列）が必要\n(b): 両試料の FASTA ファイルが必要"),
]

fill_map = {"✅ 回答可能": GREEN, "❌ 回答不可": RED}
for r_idx, (ex, q, status, reason) in enumerate(rows, 2):
    c(ws, r_idx, 1, ex,     fill=GRAY, font=bold_black, align=center)
    c(ws, r_idx, 2, q,      align=left)
    c(ws, r_idx, 3, status, fill=fill_map.get(status), font=bold_black, align=center)
    c(ws, r_idx, 4, reason, align=left)

# ===== 不足データセクション =====
sep = len(rows) + 3
title_cell = ws.cell(row=sep, column=1,
    value="【不足データ一覧 - 篠原君に提供依頼が必要なファイル】")
title_cell.font = Font(bold=True, size=11)
ws.merge_cells(start_row=sep, start_column=1, end_row=sep, end_column=4)
ws.row_dimensions[sep].height = 22

for col, h in enumerate(["種別", "ファイルパス", "内容", "提供元"], 1):
    c(ws, sep+1, col, h, fill=BLUE, font=bold_white, align=center)

needed = [
    ("Excel（例1・5用）",
     "data/02_tables__gPlant__4_相同性が高い生物種リスト(BLAST結果10位まで).xlsx",
     "r58-240918S, r26-221010M, r3-230606S 等の\n列を含む版に差替え",
     "篠原君の QIIME2 解析出力"),
    ("FASTA（例3・4用）",
     "data/04_sequences_fasta__gPlant__230606S.fasta",
     "r3-230606S の rbcL ASV 配列",
     "QIIME2 per-sample-sequences"),
    ("FASTA（例3・4用）",
     "data/04_sequences_fasta__gPlant__221010M.fasta",
     "r26-221010M の rbcL ASV 配列",
     "同上"),
    ("FASTA（例3・4用）",
     "data/04_sequences_fasta__gPlant__240918S.fasta",
     "r58-240918S の rbcL ASV 配列",
     "同上"),
]
for r_idx, row_data in enumerate(needed, sep+2):
    for col, val in enumerate(row_data, 1):
        c(ws, r_idx, col, val, fill=YELLOW, align=left)
    ws.row_dimensions[r_idx].height = 42

# ===== 列幅・行高 =====
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 56
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 48
ws.row_dimensions[1].height = 22
for r in range(2, len(rows)+2):
    ws.row_dimensions[r].height = 60
ws.row_dimensions[sep+1].height = 22

wb.save("RAG_question_status.xlsx")
print("Saved: RAG_question_status.xlsx")
