from __future__ import annotations

from pathlib import Path
import re

from data_access import find_marker_table_file, find_sequence_fasta_files, infer_marker_from_path, logical_stem
from openpyxl import load_workbook

from sample_mapping import resolve_sample_identifier, summarize_thesis_record
from task_understanding import understand_sequence_request


THESIS_SAMPLE_ID_PATTERN = r"r\d+[A-Za-z]?(?:-\d+[A-Za-z]?)+"
SEQUENCE_SAMPLE_NAME_PATTERN = r"(?<![A-Za-z0-9_])\d{6}[A-Za-z](?:-\d+)?(?![A-Za-z0-9_])"
BLAST_FILE_NAME = "4_相同性が高い生物種リスト(BLAST結果10位まで).xlsx"


def detect_sequence_sample_ids(question: str) -> list[str]:
    matches = re.findall(THESIS_SAMPLE_ID_PATTERN, question, flags=re.IGNORECASE)
    matches.extend(re.findall(SEQUENCE_SAMPLE_NAME_PATTERN, question, flags=re.IGNORECASE))
    return list(dict.fromkeys(matches))


def should_use_ncbi_blast(question: str) -> bool:
    """Return True when the question explicitly asks to use NCBI BLAST online."""
    normalized = question.lower()
    terms = [
        "blastn",
        "blast データベース",
        "blastデータベース",
        "blast db",
        "ncbi blast",
        "blast.ncbi",
        "ncbiで検索",
        "ncbiに",
        "ncbiで",
        "online blast",
    ]
    return any(term in normalized for term in terms)


def should_use_sequence_query(question: str) -> bool:
    normalized = question.lower()
    terms = [
        "fasta",
        "asv",
        "blast",
        "blastn",
        "sequence",
        "塩基配列",
        "植物種",
        "学名",
        "表",
        "割合",
        "比率",
        "top",
        "上位",
        "まとめ",
        "合計",
        "%",
    ]
    return bool(detect_sequence_sample_ids(question)) and any(term in normalized for term in terms)


def detect_output_language(question: str) -> str:
    if re.search(r"[\u3040-\u30ff]", question):
        return "ja"
    if re.search(r"[\u4e00-\u9fff]", question):
        return "zh"
    return "en"


def infer_sample_id_from_history(history: list[dict] | None) -> str | None:
    if not history:
        return None
    for item in reversed(history):
        sequence_context = item.get("sequence_context") if isinstance(item, dict) else None
        if sequence_context and sequence_context.get("sample_id"):
            return sequence_context["sample_id"]
    return None


def list_available_sequence_samples(data_dir: Path) -> list[str]:
    return [
        logical_stem(path)
        for path in find_sequence_fasta_files(data_dir)
        if logical_stem(path).lower() != "repset"
    ]


def find_fasta_for_sample(data_dir: Path, sample_name: str) -> list[Path]:
    matches = []
    for path in find_sequence_fasta_files(data_dir):
        if logical_stem(path).lower() == sample_name.lower():
            matches.append(path)
    return matches


def parse_fasta_entries(path: Path) -> list[dict]:
    entries = []
    current_header = None
    current_sequence = []

    def flush():
        if not current_header:
            return
        header = current_header[1:]
        match = re.match(r"(ASV[_\-]?\d+)_\(([\d.]+)%\)", header, flags=re.IGNORECASE)
        asv_id = match.group(1).replace("-", "_").upper() if match else header.split()[0].upper()
        ratio = float(match.group(2)) if match else None
        entries.append(
            {
                "asv_id": asv_id,
                "ratio_percent": ratio,
                "sequence": "".join(current_sequence),
            }
        )

    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith(">"):
            flush()
            current_header = line
            current_sequence = []
        else:
            current_sequence.append(line)
    flush()
    return entries


def load_blast_rows(data_dir: Path, marker: str) -> dict[str, dict]:
    path = find_marker_table_file(data_dir, marker, "4")
    if not path:
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = list(rows[0])
    top_slots = []
    slot = 1
    while f"Target_top{slot}" in header:
        top_slots.append(
            {
                "target": header.index(f"Target_top{slot}"),
                "identity": header.index(f"Identity_top{slot}"),
                "species": header.index(f"Species_top{slot}"),
            }
        )
        slot += 1

    asv_rows = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        asv_id = str(row[0]).upper()
        top_hits = []
        for slot_data in top_slots:
            target = row[slot_data["target"]] if slot_data["target"] < len(row) else None
            identity = row[slot_data["identity"]] if slot_data["identity"] < len(row) else None
            species = row[slot_data["species"]] if slot_data["species"] < len(row) else None
            if target or species:
                top_hits.append(
                    {
                        "target": str(target or ""),
                        "identity": identity,
                        "species_label": str(species or ""),
                    }
                )
        asv_rows[asv_id] = {
            "source": str(path.relative_to(data_dir)),
            "top_hits": top_hits,
        }
    return asv_rows


def build_sequence_rows(
    data_dir: Path,
    sample_name: str,
    use_ncbi: bool = False,
    blast_cache_path: Path | None = None,
) -> dict | None:
    fasta_paths = find_fasta_for_sample(data_dir, sample_name)
    if not fasta_paths:
        return None

    path = fasta_paths[0]
    marker = infer_marker_from_path(path)
    entries = parse_fasta_entries(path)

    if not use_ncbi:
        # Try local Excel first; auto-fallback to NCBI when no hits found
        blast_rows = load_blast_rows(data_dir, marker)
        has_local_hits = any(
            bool(blast_rows.get(e["asv_id"], {}).get("top_hits"))
            for e in entries
        )
        if not has_local_hits:
            use_ncbi = True  # no local data → automatic NCBI fallback

    if use_ncbi:
        from blast_ncbi import blast_entries as _blast_entries
        cache_path = blast_cache_path or (data_dir / ".blast_cache.json")
        # Submit top-20 by ratio as one batch to avoid overlong jobs
        top_entries = sorted(entries, key=lambda e: e["ratio_percent"] or 0, reverse=True)[:20]
        ncbi_results = _blast_entries(top_entries, cache_path=cache_path)
        rows = []
        for entry in entries:
            ncbi_hits = ncbi_results.get(entry["asv_id"], [])
            rows.append(
                {
                    "asv_id": entry["asv_id"],
                    "ratio_percent": entry["ratio_percent"],
                    "top_hits": [
                        {
                            "target": hit["accession"],
                            "identity": hit["identity_pct"],
                            "species_label": hit["species_label"],
                        }
                        for hit in ncbi_hits
                    ],
                }
            )
        blast_source = "NCBI BLAST online (nt database)"
    else:
        rows = []
        for entry in entries:
            blast = blast_rows.get(entry["asv_id"], {})
            rows.append(
                {
                    "asv_id": entry["asv_id"],
                    "ratio_percent": entry["ratio_percent"],
                    "top_hits": blast.get("top_hits", []),
                }
            )
        local_blast_file = find_marker_table_file(data_dir, marker, "4")
        blast_source = (
            str(local_blast_file.relative_to(data_dir))
            if local_blast_file
            else "<no-local-blast-excel>"
        )

    return {
        "sample_id": sample_name,
        "marker": marker,
        "fasta_path": str(path.relative_to(data_dir)),
        "blast_source": blast_source,
        "rows": rows,
    }


def build_grouped_species_rows(
    data_dir: Path,
    sample_name: str,
    use_ncbi: bool = False,
    blast_cache_path: Path | None = None,
) -> dict | None:
    dataset = build_sequence_rows(data_dir, sample_name, use_ncbi=use_ncbi, blast_cache_path=blast_cache_path)
    if not dataset:
        return None
    grouped = {}
    for row in dataset["rows"]:
        top1 = row["top_hits"][0]["target"] if row["top_hits"] else ""
        if not top1:
            continue
        grouped.setdefault(top1, {"asv_ids": [], "ratio_sum": 0.0})
        grouped[top1]["asv_ids"].append(row["asv_id"])
        grouped[top1]["ratio_sum"] += float(row["ratio_percent"] or 0)
    dataset["grouped_rows"] = [
        {
            "species_name": species_name,
            "asv_ids": values["asv_ids"],
            "ratio_sum": round(values["ratio_sum"], 4),
        }
        for species_name, values in sorted(grouped.items(), key=lambda item: item[1]["ratio_sum"], reverse=True)
    ]
    return dataset


def _mapping_missing_result(question: str, resolutions: list, data_dir: Path) -> dict:
    return {
        "question": question,
        "type": "sequence_mapping_missing",
        "resolutions": resolutions,
        "available_samples": list_available_sequence_samples(data_dir),
        "sequence_context": {"sample_id": resolutions[0].requested_id},
        "request_task_type": "unknown",
    }


def _sample_missing_result(question: str, sample_id: str, data_dir: Path) -> dict:
    return {
        "question": question,
        "type": "sample_missing_sequence_data",
        "sample_id": sample_id,
        "available_samples": list_available_sequence_samples(data_dir),
        "sequence_context": {"sample_id": sample_id},
        "request_task_type": "unknown",
    }


def answer_sequence_query(question: str, data_dir: Path, history: list[dict] | None = None, chat_model=None) -> dict | None:
    use_ncbi = should_use_ncbi_blast(question)
    request = understand_sequence_request(question, chat_model=chat_model)
    if not should_use_sequence_query(question) and request["task_type"] == "unknown":
        return None

    sample_ids = detect_sequence_sample_ids(question)
    if not sample_ids and history and request["use_previous_context"]:
        inferred = infer_sample_id_from_history(history)
        if inferred:
            sample_ids = [inferred]

    if not sample_ids:
        if should_use_sequence_query(question) or request["task_type"] != "unknown":
            inferred = infer_sample_id_from_history(history)
            if inferred:
                sample_ids = [inferred]
            else:
                return None
        else:
            return None

    if request["task_type"] == "compare_species_sets":
        if len(sample_ids) < 2:
            inferred = infer_sample_id_from_history(history)
            if inferred and inferred not in sample_ids:
                sample_ids.append(inferred)
        if len(sample_ids) < 2:
            return None

        unresolved = []
        missing = []
        datasets = []
        for sample_id in sample_ids[:2]:
            resolution = resolve_sample_identifier(data_dir, sample_id)
            if resolution.status == "mapping_missing":
                unresolved.append(resolution)
                continue
            if resolution.status == "missing":
                missing.append(sample_id)
                continue

            resolved_name = resolution.resolved_sample_name or sample_id
            dataset = build_grouped_species_rows(data_dir, resolved_name, use_ncbi=use_ncbi)
            if not dataset:
                missing.append(sample_id)
                continue
            dataset["requested_sample_id"] = sample_id
            dataset["resolved_sample_name"] = resolved_name
            datasets.append(dataset)

        if unresolved:
            result = _mapping_missing_result(question, unresolved, data_dir)
            result["request_task_type"] = request["task_type"]
            return result
        if missing:
            result = _sample_missing_result(question, ", ".join(missing), data_dir)
            result["request_task_type"] = request["task_type"]
            return result

        left, right = datasets[0], datasets[1]
        left_species = {row["species_name"]: row for row in left["grouped_rows"]}
        right_species = {row["species_name"]: row for row in right["grouped_rows"]}
        shared = sorted(set(left_species) & set(right_species))
        only_left = sorted(set(left_species) - set(right_species))
        only_right = sorted(set(right_species) - set(left_species))
        return {
            "question": question,
            "type": "compare_species_sets",
            "left": left,
            "right": right,
            "shared_species": shared,
            "only_left_species": only_left,
            "only_right_species": only_right,
            "sequence_context": {"sample_id": sample_ids[0]},
        }

    sample_id = sample_ids[0]
    resolution = resolve_sample_identifier(data_dir, sample_id)
    if resolution.status == "mapping_missing":
        result = _mapping_missing_result(question, [resolution], data_dir)
        result["request_task_type"] = request["task_type"]
        return result
    if resolution.status == "missing":
        result = _sample_missing_result(question, sample_id, data_dir)
        result["request_task_type"] = request["task_type"]
        return result

    resolved_name = resolution.resolved_sample_name or sample_id
    fasta_paths = find_fasta_for_sample(data_dir, resolved_name)
    if not fasta_paths:
        result = _sample_missing_result(question, sample_id, data_dir)
        result["request_task_type"] = request["task_type"]
        return result

    if request["task_type"] == "grouped_species_table":
        dataset = build_grouped_species_rows(data_dir, resolved_name, use_ncbi=use_ncbi)
        return {
            "question": question,
            "type": "blast_grouped_species_table",
            **dataset,
            "requested_sample_id": sample_id,
            "resolved_sample_name": resolved_name,
            "sequence_context": {"sample_id": sample_id},
        }

    if request["task_type"] == "blast_top1_table":
        dataset = build_sequence_rows(data_dir, resolved_name, use_ncbi=use_ncbi)
        return {
            "question": question,
            "type": "blast_top1_table",
            **dataset,
            "requested_sample_id": sample_id,
            "resolved_sample_name": resolved_name,
            "sequence_context": {"sample_id": sample_id},
            "response_requirements": request,
        }

    if request["task_type"] == "blast_top10_table":
        dataset = build_sequence_rows(data_dir, resolved_name, use_ncbi=use_ncbi)
        return {
            "question": question,
            "type": "blast_table",
            **dataset,
            "requested_sample_id": sample_id,
            "resolved_sample_name": resolved_name,
            "sequence_context": {"sample_id": sample_id},
            "response_requirements": request,
        }

    if request["task_type"] == "fasta_count":
        files = []
        total_entries = 0
        for path in fasta_paths:
            entries = parse_fasta_entries(path)
            files.append({"path": str(path.relative_to(data_dir)), "entry_count": len(entries)})
            total_entries += len(entries)
        return {
            "question": question,
            "type": "fasta_count",
            "sample_id": sample_id,
            "resolved_sample_name": resolved_name,
            "files": files,
            "total_entries": total_entries,
            "sequence_context": {"sample_id": sample_id},
            "response_requirements": request,
        }

    return None


# ---------------------------------------------------------------------------
# Gather pattern: for any sample id detected in the question (or carried over
# from history), pull whatever FASTA / BLAST data the local files hold and
# return it as a structured payload. No keyword gates.
# ---------------------------------------------------------------------------


def detect_fasta_sample_names(question: str, data_dir: Path) -> list[str]:
    """Return available FASTA sample names that appear literally in the question."""
    available = list_available_sequence_samples(data_dir)
    q_lower = question.lower()
    return [name for name in available if name.lower() in q_lower]


def gather_sequence_data(
    question: str,
    data_dir: Path,
    history: list[dict] | None = None,
    chat_model=None,
) -> dict | None:
    sample_ids = detect_sequence_sample_ids(question)
    if not sample_ids and history:
        inferred = infer_sample_id_from_history(history)
        if inferred:
            sample_ids = [inferred]
    # Fallback: match literal FASTA sample names present in the question
    if not sample_ids:
        sample_ids = detect_fasta_sample_names(question, data_dir)
    if not sample_ids:
        return None

    # Explicit "blastn"/"ncbi" keyword forces NCBI even when local data exists.
    # Without the keyword, build_sequence_rows auto-falls back to NCBI when local has no hits.
    use_ncbi = should_use_ncbi_blast(question)
    available_set = set(s.lower() for s in list_available_sequence_samples(data_dir))
    samples = []
    for sample_id in sample_ids:
        # If the sample_id directly matches an available FASTA file, skip resolver
        if sample_id.lower() in available_set:
            dataset = build_grouped_species_rows(data_dir, sample_id, use_ncbi=use_ncbi)
            info = {
                "requested_id": sample_id,
                "resolution_status": "resolved",
                "resolved_sample_name": sample_id,
                "thesis_token": None,
                "thesis_record": None,
                "evidence": ["direct FASTA match"],
                "dataset": dataset,
                "ncbi_blast_used": use_ncbi,
            }
            samples.append(info)
            continue

        resolution = resolve_sample_identifier(data_dir, sample_id)
        info = {
            "requested_id": sample_id,
            "resolution_status": resolution.status,
            "resolved_sample_name": resolution.resolved_sample_name,
            "thesis_token": resolution.thesis_token,
            "thesis_record": summarize_thesis_record(resolution.thesis_record),
            "evidence": list(resolution.evidence),
            "dataset": None,
            "ncbi_blast_used": use_ncbi,
        }
        if resolution.status not in {"missing", "mapping_missing"}:
            resolved_name = resolution.resolved_sample_name or sample_id
            dataset = build_grouped_species_rows(data_dir, resolved_name, use_ncbi=use_ncbi)
            if dataset:
                info["dataset"] = dataset
        samples.append(info)

    return {
        "sample_ids": sample_ids,
        "available_samples": list_available_sequence_samples(data_dir),
        "samples": samples,
        "ncbi_blast_used": use_ncbi,
    }


def format_sequence_entries(gathered) -> list[dict]:
    if not gathered:
        return []
    entries = []
    available = gathered["available_samples"]
    for sample in gathered["samples"]:
        lines = [
            f"Requested sample: {sample['requested_id']}",
            f"Resolution status: {sample['resolution_status']}",
        ]
        if sample.get("resolved_sample_name"):
            lines.append(f"Resolved sequence sample name: {sample['resolved_sample_name']}")
        if sample.get("thesis_token"):
            lines.append(f"Thesis token: {sample['thesis_token']}")
        if sample.get("thesis_record"):
            lines.append(f"Thesis record: {sample['thesis_record']}")
        if sample.get("evidence"):
            lines.append(f"Resolution evidence: {sample['evidence']}")

        use_ncbi = sample.get("ncbi_blast_used", False)
        dataset = sample.get("dataset")
        source = f"<no-fasta-for-{sample['requested_id']}>"
        if dataset:
            source = dataset["fasta_path"]
            lines.append(f"FASTA path: {dataset['fasta_path']}")
            lines.append(f"BLAST source: {dataset['blast_source']}")
            if use_ncbi:
                lines.append("NOTE: BLAST results retrieved from NCBI BLAST online (nt database)")
            lines.append(f"ASV count: {len(dataset['rows'])}")
            for row in dataset["rows"][:30]:
                top1_target = row["top_hits"][0]["target"] if row["top_hits"] else ""
                top1_species = row["top_hits"][0].get("species_label", "") if row["top_hits"] else ""
                lines.append(
                    f"  - {row['asv_id']}: ratio={row['ratio_percent']}, "
                    f"top1_target={top1_target}, top1_species={top1_species}"
                )
            grouped = dataset.get("grouped_rows") or []
            if grouped:
                lines.append(f"Aggregated by top1 species ({len(grouped)} groups):")
                for group in grouped:
                    lines.append(
                        f"  - {group['species_name']}: ratio_sum={group['ratio_sum']}%, "
                        f"asvs={group['asv_ids']}"
                    )
        else:
            status = sample.get("resolution_status", "unknown")
            if status == "mapping_missing":
                token = sample.get("thesis_token") or ""
                lines.append(
                    f"MAPPING MISSING: sample '{sample['requested_id']}' exists in the thesis "
                    f"but has no corresponding FASTA file (thesis token='{token}' "
                    f"does not match any file in data/04_sequences_fasta/)."
                )
                lines.append(
                    "To fix: add the per-sample FASTA file (e.g. "
                    f"'data/04_sequences_fasta__gPlant__{token}.fasta') "
                    "OR populate data/sample_id_mapping.csv with the correct mapping."
                )
            else:
                lines.append(
                    f"NO SEQUENCE DATA: sample '{sample['requested_id']}' was not found "
                    "in the available sequence files."
                )
        lines.append(f"Available sequence samples on disk: {available}")

        entries.append({"kind": "sequence", "source": source, "text": "\n".join(lines)})
    return entries


def _display_sample_label(result: dict, language: str) -> str:
    requested = result.get("requested_sample_id") or result.get("sample_id")
    resolved = result.get("resolved_sample_name") or result.get("sample_id")
    if requested and requested != resolved:
        if language == "ja":
            return f"{requested}（sequence サンプル名: {resolved}）"
        if language == "zh":
            return f"{requested}（对应 sequence 样本名：{resolved}）"
        return f"{requested} (sequence sample: {resolved})"
    return resolved or requested or ""


def format_sequence_answer(result: dict) -> str:
    language = detect_output_language(result.get("question", ""))

    if result["type"] == "sequence_mapping_missing":
        details = []
        for resolution in result["resolutions"]:
            metadata = summarize_thesis_record(resolution.thesis_record)
            line = resolution.requested_id
            if resolution.thesis_token:
                line += f" [token={resolution.thesis_token}]"
            if metadata:
                line += f" ({metadata})"
            if resolution.evidence:
                line += f"; evidence={', '.join(resolution.evidence)}"
            details.append(line)

        available = ", ".join(result["available_samples"])
        request_task_type = result.get("request_task_type")
        if language == "ja":
            if request_task_type == "fasta_count":
                lines = [
                    "現時点のローカルデータだけでは、指定された試料番号に含まれる fasta ファイル数は確定できません。",
                    "理由は、修士論文の試料番号と sequence サンプル名を結ぶ対応表が現在の作業ディレクトリにないためです。",
                ]
            else:
                lines = [
                    "試料番号は修士論文側では確認できましたが、現在のローカル sequence / BLAST データには対応表がありません。",
                    "そのため、この質問に対して信頼できる fasta / ASV / BLAST 結果を確定できません。",
                ]
            lines.extend(f"- {item}" for item in details)
            lines.append(f"現在参照可能な sequence サンプル: {available}")
            return "\n".join(lines)
        if language == "zh":
            if request_task_type == "fasta_count":
                lines = [
                    "仅根据当前本地数据，无法确定该试料编号对应的 fasta 文件数量。",
                    "原因是当前工作目录里没有“论文试料编号 -> sequence 样本名”的对应表。",
                ]
            else:
                lines = [
                    "该试料编号在论文样本信息中可以确认，但当前本地 sequence / BLAST 数据里没有可靠的对应映射。",
                    "因此现在不能把它稳定地对应到某个 fasta / ASV / BLAST 结果上。",
                ]
            lines.extend(f"- {item}" for item in details)
            lines.append(f"当前可用的 sequence 样本有：{available}")
            return "\n".join(lines)
        lines = [
            "The thesis sample ID exists, but there is no reliable mapping to the local sequence / BLAST sample names.",
            "So a fasta / ASV / BLAST answer cannot be determined safely from the current data.",
        ]
        lines.extend(f"- {item}" for item in details)
        lines.append(f"Available sequence samples: {available}")
        return "\n".join(lines)

    if result["type"] == "sample_missing_sequence_data":
        sample_id = result["sample_id"]
        available = ", ".join(result["available_samples"])
        if language == "ja":
            return (
                f"{sample_id} に対応する fasta / BLAST データは、現在のローカルデータには見つかりませんでした。\n"
                f"現在参照可能な sequence サンプル: {available}"
            )
        if language == "zh":
            return (
                f"当前本地数据中没有找到与 {sample_id} 对应的 fasta / BLAST 数据。\n"
                f"目前可用的 sequence 样本有：{available}"
            )
        return (
            f"No local fasta / BLAST data was found for {sample_id}.\n"
            f"Available sequence samples: {available}"
        )

    if result["type"] == "fasta_count":
        label = _display_sample_label(result, language)
        if language == "ja":
            lines = [f"{label} に対応する fasta 配列数は {result['total_entries']} です。"]
        elif language == "zh":
            lines = [f"{label} 对应的 fasta 序列数为 {result['total_entries']}。"]
        else:
            lines = [f"{label} has {result['total_entries']} fasta entries."]
        for item in result["files"]:
            lines.append(f"- {item['path']}: {item['entry_count']}")
        return "\n".join(lines)

    if result["type"] == "blast_table":
        label = _display_sample_label(result, language)
        if language == "ja":
            header = f"{label} の ASV と、それぞれに対応する植物種（学名）の上位10位です。"
            col_headers = "| ASV番号 | リード数の割合（%） | 植物種（学名）上位10位 |"
        elif language == "zh":
            header = f"{label} 各 ASV 对应的生物物种（学名）前10位。"
            col_headers = "| ASV 编号 | reads 占比（%） | 生物物种（学名）前10位 |"
        else:
            header = f"Top-10 BLAST species for each ASV in {label}."
            col_headers = "| ASV ID | Reads ratio (%) | Top-10 species (scientific name) |"
        lines = [
            header,
            f"fasta: {result['fasta_path']}",
            f"BLAST source: {result['blast_source']}",
            "",
            col_headers,
            "|---|---:|---|",
        ]
        for row in result["rows"]:
            species = [hit["target"] for hit in row["top_hits"][:10]]
            lines.append(
                f"| {row['asv_id']} | {row['ratio_percent'] if row['ratio_percent'] is not None else ''} | "
                f"{'<br>'.join(species)} |"
            )
        return "\n".join(lines)

    if result["type"] == "blast_top1_table":
        label = _display_sample_label(result, language)
        if language == "ja":
            header = f"{label} の ASV と、それぞれに対応する植物種（学名）の上位1位です。"
            col_headers = "| ASV番号 | リード数の割合（%） | 植物種（学名）上位1位 |"
        elif language == "zh":
            header = f"{label} 各 ASV 对应的生物物种（学名）第1位。"
            col_headers = "| ASV 编号 | reads 占比（%） | 生物物种（学名）第1位 |"
        else:
            header = f"Top-1 BLAST species for each ASV in {label}."
            col_headers = "| ASV ID | Reads ratio (%) | Top-1 species (scientific name) |"
        lines = [
            header,
            f"fasta: {result['fasta_path']}",
            f"BLAST source: {result['blast_source']}",
            "",
            col_headers,
            "|---|---:|---|",
        ]
        for row in result["rows"]:
            top1 = row["top_hits"][0]["target"] if row["top_hits"] else ""
            lines.append(
                f"| {row['asv_id']} | {row['ratio_percent'] if row['ratio_percent'] is not None else ''} | {top1} |"
            )
        return "\n".join(lines)

    if result["type"] == "blast_grouped_species_table":
        label = _display_sample_label(result, language)
        if language == "ja":
            header = f"{label} について、同じ植物種（学名）ごとにまとめた表です。"
            col_headers = "| 植物種（学名） | 含まれるASV番号 | リード数の合計値（%） |"
        elif language == "zh":
            header = f"{label} 按生物物种（学名）汇总的表格。"
            col_headers = "| 生物物种（学名） | 包含的 ASV 编号 | reads 占比合计（%） |"
        else:
            header = f"Species-grouped BLAST table for {label}."
            col_headers = "| Species (scientific name) | ASV IDs | Reads ratio sum (%) |"
        lines = [
            header,
            f"fasta: {result['fasta_path']}",
            f"BLAST source: {result['blast_source']}",
            "",
            col_headers,
            "|---|---|---:|",
        ]
        for row in result["grouped_rows"]:
            lines.append(f"| {row['species_name']} | {', '.join(row['asv_ids'])} | {row['ratio_sum']} |")
        return "\n".join(lines)

    if result["type"] == "compare_species_sets":
        left_name = _display_sample_label(result["left"], language)
        right_name = _display_sample_label(result["right"], language)
        if language == "ja":
            lines = [f"{left_name} と {right_name} の比較です。", ""]
            lines.append("共通して見られた植物種（学名）:")
            lines.extend([f"- {name}" for name in result["shared_species"]] or ["- なし"])
            lines.append("")
            lines.append(f"{left_name} のみに見られた植物種（学名）:")
            lines.extend([f"- {name}" for name in result["only_left_species"]] or ["- なし"])
            lines.append("")
            lines.append(f"{right_name} のみに見られた植物種（学名）:")
            lines.extend([f"- {name}" for name in result["only_right_species"]] or ["- なし"])
        elif language == "zh":
            lines = [f"{left_name} 与 {right_name} 的比较。", ""]
            lines.append("共同检出的生物物种（学名）：")
            lines.extend([f"- {name}" for name in result["shared_species"]] or ["- 无"])
            lines.append("")
            lines.append(f"仅在 {left_name} 中检出的生物物种（学名）：")
            lines.extend([f"- {name}" for name in result["only_left_species"]] or ["- 无"])
            lines.append("")
            lines.append(f"仅在 {right_name} 中检出的生物物种（学名）：")
            lines.extend([f"- {name}" for name in result["only_right_species"]] or ["- 无"])
        else:
            lines = [f"Comparison: {left_name} vs {right_name}.", ""]
            lines.append("Species found in both:")
            lines.extend([f"- {name}" for name in result["shared_species"]] or ["- (none)"])
            lines.append("")
            lines.append(f"Species only in {left_name}:")
            lines.extend([f"- {name}" for name in result["only_left_species"]] or ["- (none)"])
            lines.append("")
            lines.append(f"Species only in {right_name}:")
            lines.extend([f"- {name}" for name in result["only_right_species"]] or ["- (none)"])
        return "\n".join(lines)

    return ""
