from pathlib import Path
import re

from data_access import find_marker_table_file
from openpyxl import load_workbook

from query_intent import infer_table_query_intent
from sample_mapping import resolve_sample_identifier, summarize_thesis_record
from task_understanding import understand_table_request


RESULT_FILE_NAMES = {
    "3": "3_代表配列のリード数と相同性解析結果.xlsx",
    "4": "4_相同性が高い生物種リスト(BLAST結果10位まで).xlsx",
}


def parse_species_name(target: object) -> str:
    if not target:
        return "Unknown"

    text = str(target)
    text = re.sub(r"^[A-Z]{1,3}_?\d+(?:\.\d+)?_", "", text)
    text = text.replace(",", " ").replace(";", " ")
    parts = text.split("_")

    for index in range(len(parts) - 1):
        genus = parts[index]
        species = parts[index + 1]
        if genus and species and genus[0].isupper() and species.islower():
            return f"{genus} {species}"

    return text[:100]


def detect_markers(question: str) -> list[str]:
    normalized = question.lower()
    markers = []
    if "coi" in normalized:
        markers.append("COI")
    if any(term in normalized for term in ["gplant", "rbcl", "rbc"]):
        markers.append("gPlant")
    return markers


def infer_markers(question: str, llm_markers: list[str] | None = None) -> list[str]:
    if llm_markers:
        return ["COI", "gPlant"] if "all" in llm_markers else llm_markers
    markers = detect_markers(question)
    return markers or ["COI", "gPlant"]


def detect_asv_ids(question: str) -> list[str]:
    matches = re.findall(r"ASV[_\-\s]?(\d+)", question, flags=re.IGNORECASE)
    return list(dict.fromkeys(f"ASV_{int(match):03d}" for match in matches))


def detect_sample_ids(question: str) -> list[str]:
    patterns = [
        r"(?<![A-Za-z0-9_])r\d+-\d+[A-Za-z]?(?![A-Za-z0-9_])",
        r"(?<![A-Za-z0-9_])\d{6}[A-Za-z](?:-\d+)?(?![A-Za-z0-9_])",
    ]
    matches = []
    for pattern in patterns:
        matches.extend(re.findall(pattern, question, flags=re.IGNORECASE))
    return list(dict.fromkeys(matches))


def detect_table_file_preference(question: str) -> str:
    normalized = question.lower()
    if any(term in normalized for term in ["4_", "blast結果10位まで", "相同性が高い生物種リスト", "top10", "top 10"]):
        return "4"
    if any(term in normalized for term in ["3_", "代表配列", "相同性解析結果"]):
        return "3"
    return "3"


def detect_species_name(question: str) -> str | None:
    match = re.search(r"\b([A-Z][a-z]+)\s+([a-z][a-z.\-]+)\b", question)
    if not match:
        return None
    return f"{match.group(1)} {match.group(2)}"


def is_context_dependent_table_question(question: str) -> bool:
    normalized = question.lower().strip()
    context_terms = ["this", "that", "it", "those", "上面", "刚才", "前面", "这个", "那个", "これ", "それ", "前の"]
    return len(normalized) <= 20 or any(term in normalized for term in context_terms)


ALL_LIMIT = 9999

_ALL_TERMS = ["すべて", "全部", "全て", "全種", "全asvを", "全asv", "all asv", "all species"]


def detect_limit(question: str, default: int = 10) -> int:
    normalized = question.lower()
    if any(term in normalized for term in _ALL_TERMS):
        return ALL_LIMIT
    patterns = [
        r"top\s*(\d+)",
        r"upper\s*(\d+)",
        r"前\s*(\d+)",
        r"上位\s*(\d+)",
        r"(\d+)\s*件",
        r"(\d+)\s*个",
        r"(\d+)\s*種",
    ]
    for pattern in patterns:
        match = re.search(pattern, question, flags=re.IGNORECASE)
        if match:
            return max(1, min(int(match.group(1)), 50))
    return default


def _heuristic_table_intent(question: str) -> dict:
    asv_ids = detect_asv_ids(question)
    sample_ids = detect_sample_ids(question)
    species_name = detect_species_name(question)
    normalized = question.lower()

    if asv_ids:
        query_type = "asv_details"
    elif sample_ids and any(term in normalized for term in ["合計", "sum", "total", "総", "リード数の合計", "reads total"]):
        query_type = "sample_total_reads"
    elif species_name:
        query_type = "species_matches"
    elif sample_ids:
        query_type = "sample_asvs" if "asv" in normalized else "sample_species"
    elif any(term in normalized for term in ["asv", "otu"]):
        query_type = "top_asvs"
    else:
        query_type = "top_species"

    should_use = bool(asv_ids or sample_ids or species_name)
    if not should_use:
        broad_terms = [
            "read",
            "reads",
            "blast",
            "identity",
            "species",
            "detected",
            "sample",
            "top",
            "most",
            "abundance",
            "物种",
            "生物",
            "检出",
            "配列",
            "相同性",
        ]
        should_use = any(term in normalized for term in broad_terms)

    return {
        "should_use_table_query": should_use,
        "query_type": query_type,
        "markers": infer_markers(question),
        "asv_ids": asv_ids,
        "sample_ids": sample_ids,
        "species_name": species_name,
        "limit": detect_limit(question),
        "table_file": detect_table_file_preference(question),
    }


def detect_explicit_query_type(question: str, sample_ids: list[str], asv_ids: list[str]) -> str | None:
    normalized = question.lower()
    if sample_ids and any(term in normalized for term in ["合計", "sum", "total", "総", "リード数の合計", "reads total"]):
        return "sample_total_reads"
    if sample_ids and any(term in normalized for term in ["asv", "otu"]):
        return "sample_asvs"
    if asv_ids:
        return "asv_details"
    return None


def should_use_table_query(question: str, chat_model=None) -> bool:
    if is_context_dependent_table_question(question) and not (
        detect_asv_ids(question) or detect_sample_ids(question) or detect_species_name(question)
    ):
        return False

    intent = infer_table_query_intent(question, chat_model)
    if intent:
        return bool(intent["should_use_table_query"])
    return bool(_heuristic_table_intent(question)["should_use_table_query"])


def load_result_rows(data_dir: Path, marker: str, table_file: str = "3") -> list[dict]:
    path = find_marker_table_file(data_dir, marker, table_file)
    if not path:
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = list(rows[0])
    target_column_name = "Target" if "Target" in header else "Target_top1"
    identity_column_name = "Identity" if "Identity" in header else "Identity_top1"
    target_index = header.index(target_column_name)
    identity_index = header.index(identity_column_name)
    sample_columns = [
        (index, str(value))
        for index, value in enumerate(header[:target_index])
        if index > 0 and value
    ]

    records = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue

        sample_reads = {}
        total_reads = 0
        for index, sample_name in sample_columns:
            value = row[index] if index < len(row) else 0
            try:
                read_count = int(value or 0)
            except (TypeError, ValueError):
                read_count = 0
            sample_reads[sample_name] = read_count
            total_reads += read_count

        target = row[target_index] if target_index < len(row) else ""
        identity = row[identity_index] if identity_index < len(row) else ""
        records.append(
            {
                "marker": marker,
                "asv_id": str(row[0]),
                "sample_reads": sample_reads,
                "total_reads": total_reads,
                "target": target,
                "species": parse_species_name(target),
                "identity": identity,
                "source": str(path.relative_to(data_dir)),
                "table_file": table_file,
            }
        )

    return records


def get_top_species(data_dir: Path, marker: str, limit: int = 10, table_file: str = "3") -> list[dict]:
    totals = {}
    examples = {}
    for row in load_result_rows(data_dir, marker, table_file=table_file):
        species = row["species"]
        totals[species] = totals.get(species, 0) + row["total_reads"]
        examples.setdefault(species, row)

    return [
        {
            "marker": marker,
            "species": species,
            "total_reads": total_reads,
            "example_asv": examples[species]["asv_id"],
            "source": examples[species]["source"],
        }
        for species, total_reads in sorted(totals.items(), key=lambda item: item[1], reverse=True)[:limit]
    ]


def table_has_sample(data_dir: Path, marker: str, sample_id: str, table_file: str = "3") -> bool:
    return any(sample_id in row["sample_reads"] for row in load_result_rows(data_dir, marker, table_file=table_file))


def get_top_asvs(data_dir: Path, marker: str, limit: int = 10, table_file: str = "3") -> list[dict]:
    rows = load_result_rows(data_dir, marker, table_file=table_file)
    return sorted(rows, key=lambda row: row["total_reads"], reverse=True)[:limit]


def get_asv_detail(data_dir: Path, marker: str, asv_id: str, table_file: str = "3") -> dict | None:
    for row in load_result_rows(data_dir, marker, table_file=table_file):
        if row["asv_id"].upper() == asv_id.upper():
            return row
    return None


def find_species(data_dir: Path, marker: str, species_name: str, table_file: str = "3") -> list[dict]:
    normalized = species_name.lower()
    rows = load_result_rows(data_dir, marker, table_file=table_file)
    return [
        row for row in rows
        if normalized in row["species"].lower() or normalized in str(row["target"]).lower()
    ]


def get_sample_species(data_dir: Path, marker: str, sample_id: str, limit: int = 10, table_file: str = "3") -> list[dict]:
    totals = {}
    examples = {}
    for row in load_result_rows(data_dir, marker, table_file=table_file):
        sample_reads = int(row["sample_reads"].get(sample_id, 0) or 0)
        if sample_reads <= 0:
            continue
        species = row["species"]
        totals[species] = totals.get(species, 0) + sample_reads
        examples.setdefault(species, row)

    sorted_items = sorted(totals.items(), key=lambda item: item[1], reverse=True)
    if limit < ALL_LIMIT:
        sorted_items = sorted_items[:limit]
    return [
        {
            "marker": marker,
            "sample_id": sample_id,
            "species": species,
            "sample_reads": sample_reads,
            "example_asv": examples[species]["asv_id"],
            "source": examples[species]["source"],
        }
        for species, sample_reads in sorted_items
    ]


def get_sample_asvs(data_dir: Path, marker: str, sample_id: str, limit: int = 10, table_file: str = "3") -> list[dict]:
    rows = []
    for row in load_result_rows(data_dir, marker, table_file=table_file):
        sample_reads = int(row["sample_reads"].get(sample_id, 0) or 0)
        if sample_reads <= 0:
            continue
        enriched = dict(row)
        enriched["sample_id"] = sample_id
        enriched["sample_read_count"] = sample_reads
        rows.append(enriched)
    sorted_rows = sorted(rows, key=lambda row: row["sample_read_count"], reverse=True)
    return sorted_rows if limit >= ALL_LIMIT else sorted_rows[:limit]


def get_sample_total_reads(data_dir: Path, marker: str, sample_id: str, table_file: str = "3") -> dict | None:
    total_reads = 0
    source = None
    for row in load_result_rows(data_dir, marker, table_file=table_file):
        sample_reads = int(row["sample_reads"].get(sample_id, 0) or 0)
        total_reads += sample_reads
        source = source or row["source"]
    if source is None:
        return None
    return {
        "marker": marker,
        "sample_id": sample_id,
        "total_reads": total_reads,
        "source": source,
        "table_file": table_file,
    }


def answer_table_query(question: str, data_dir: Path, chat_model=None) -> dict | None:
    if is_context_dependent_table_question(question) and not (
        detect_asv_ids(question) or detect_sample_ids(question) or detect_species_name(question)
    ):
        return None

    llm_intent = infer_table_query_intent(question, chat_model)
    heuristic_intent = _heuristic_table_intent(question)
    intent = llm_intent or heuristic_intent

    if not intent["should_use_table_query"]:
        return None

    markers = infer_markers(question, intent.get("markers"))
    asv_ids = intent.get("asv_ids") or heuristic_intent["asv_ids"]
    sample_ids = intent.get("sample_ids") or heuristic_intent["sample_ids"]
    species_name = intent.get("species_name") or heuristic_intent["species_name"]
    heuristic_limit = heuristic_intent["limit"]
    limit = ALL_LIMIT if heuristic_limit >= ALL_LIMIT else int(intent.get("limit") or heuristic_limit or 10)
    query_type = intent.get("query_type") or heuristic_intent["query_type"]
    table_file = heuristic_intent.get("table_file", "3")
    explicit_query_type = detect_explicit_query_type(question, sample_ids, asv_ids)
    if explicit_query_type:
        query_type = explicit_query_type
    results = []

    for marker in markers:
        source_path = find_marker_table_file(data_dir, marker, table_file)
        missing_sample_ids = []
        mapping_missing = []
        resolved_sample_ids = []
        if sample_ids:
            for sample_id in sample_ids:
                resolution = resolve_sample_identifier(data_dir, sample_id)
                if resolution.status == "mapping_missing":
                    mapping_missing.append(resolution)
                    continue
                resolved_name = resolution.resolved_sample_name or sample_id
                if not table_has_sample(data_dir, marker, resolved_name, table_file=table_file):
                    missing_sample_ids.append(sample_id)
                    continue
                resolved_sample_ids.append((sample_id, resolved_name))

            if mapping_missing and not resolved_sample_ids:
                results.append(
                    {
                        "type": "sequence_mapping_missing",
                        "marker": marker,
                        "resolutions": mapping_missing,
                        "source": str(source_path.relative_to(data_dir)) if source_path else "",
                    }
                )
                continue

            if not resolved_sample_ids and len(missing_sample_ids) == len(sample_ids):
                results.append(
                    {
                        "type": "sample_not_found",
                        "marker": marker,
                        "sample_ids": missing_sample_ids,
                        "table_file": table_file,
                        "source": str(source_path.relative_to(data_dir)) if source_path else "",
                    }
                )
                continue

        if query_type == "asv_details" and asv_ids:
            details = [
                detail for asv_id in asv_ids
                if (detail := get_asv_detail(data_dir, marker, asv_id, table_file=table_file))
            ]
            if details:
                results.append({"type": "asv_details", "marker": marker, "data": details})
            continue

        if query_type == "species_matches" and species_name:
            matches = find_species(data_dir, marker, species_name, table_file=table_file)
            if matches:
                results.append(
                    {
                        "type": "species_matches",
                        "marker": marker,
                        "species_name": species_name,
                        "data": sorted(matches, key=lambda row: row["total_reads"], reverse=True)[:limit],
                    }
                )
            continue

        if query_type == "sample_asvs" and sample_ids:
            for requested_sample_id, resolved_sample_id in resolved_sample_ids:
                data = get_sample_asvs(data_dir, marker, resolved_sample_id, limit, table_file=table_file)
                if data:
                    results.append(
                        {
                            "type": "sample_asvs",
                            "marker": marker,
                            "sample_id": requested_sample_id,
                            "resolved_sample_name": resolved_sample_id,
                            "data": data,
                        }
                    )
            continue

        if query_type == "sample_species" and sample_ids:
            for requested_sample_id, resolved_sample_id in resolved_sample_ids:
                data = get_sample_species(data_dir, marker, resolved_sample_id, limit, table_file=table_file)
                if data:
                    results.append(
                        {
                            "type": "sample_species",
                            "marker": marker,
                            "sample_id": requested_sample_id,
                            "resolved_sample_name": resolved_sample_id,
                            "data": data,
                        }
                    )
            continue

        if query_type == "sample_total_reads" and sample_ids:
            for requested_sample_id, resolved_sample_id in resolved_sample_ids:
                data = get_sample_total_reads(data_dir, marker, resolved_sample_id, table_file=table_file)
                if data:
                    results.append(
                        {
                            "type": "sample_total_reads",
                            "marker": marker,
                            "sample_id": requested_sample_id,
                            "resolved_sample_name": resolved_sample_id,
                            "data": data,
                        }
                    )
            continue

        if query_type == "top_asvs":
            results.append({"type": "top_asvs", "marker": marker, "data": get_top_asvs(data_dir, marker, limit, table_file=table_file)})
            continue

        results.append({"type": "top_species", "marker": marker, "data": get_top_species(data_dir, marker, limit, table_file=table_file)})

    if not results:
        return None

    return {
        "question": question,
        "results": results,
        "response_requirements": understand_table_request(question, chat_model=chat_model),
    }


# ---------------------------------------------------------------------------
# Gather pattern: always returns structured data when something can be fetched.
# Does NOT decide whether to "answer" the question — composer LLM does that.
# ---------------------------------------------------------------------------


def load_table_columns(data_dir, marker: str, table_file: str = "3"):
    path = find_marker_table_file(data_dir, marker, table_file)
    if not path:
        return [], None
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], str(path.relative_to(data_dir))
    header = list(rows[0])
    target_column_name = "Target" if "Target" in header else "Target_top1"
    if target_column_name not in header:
        return [], str(path.relative_to(data_dir))
    target_index = header.index(target_column_name)
    columns = [
        str(value)
        for index, value in enumerate(header[:target_index])
        if index > 0 and value
    ]
    return columns, str(path.relative_to(data_dir))


def _resolve_sample_status(data_dir, sample_id: str, columns: list[str]) -> dict:
    resolution = resolve_sample_identifier(data_dir, sample_id)
    resolved_name = resolution.resolved_sample_name or sample_id
    actual_column = None
    for candidate in (resolved_name, resolution.thesis_token, sample_id):
        if candidate and candidate in columns:
            actual_column = candidate
            break
    return {
        "requested_id": sample_id,
        "resolution_status": resolution.status,
        "resolved_sample_name": resolution.resolved_sample_name,
        "thesis_token": resolution.thesis_token,
        "thesis_record": summarize_thesis_record(resolution.thesis_record),
        "evidence": list(resolution.evidence),
        "present_as_column": actual_column is not None,
        "actual_column": actual_column,
    }


def gather_table_data(question: str, data_dir, chat_model=None) -> dict | None:
    asv_ids = detect_asv_ids(question)
    sample_ids = detect_sample_ids(question)
    species_name = detect_species_name(question)

    llm_intent = infer_table_query_intent(question, chat_model)
    heuristic_intent = _heuristic_table_intent(question)
    intent = llm_intent or heuristic_intent

    has_entity = bool(asv_ids or sample_ids or species_name)
    suggested_use = bool(intent.get("should_use_table_query")) or has_entity
    if not suggested_use:
        return None

    markers = infer_markers(question, intent.get("markers"))
    asv_ids = intent.get("asv_ids") or asv_ids
    sample_ids = intent.get("sample_ids") or sample_ids
    species_name = intent.get("species_name") or species_name
    heuristic_limit = heuristic_intent["limit"]
    limit = ALL_LIMIT if heuristic_limit >= ALL_LIMIT else int(intent.get("limit") or heuristic_limit or 10)
    query_type = intent.get("query_type") or heuristic_intent["query_type"]
    table_file = heuristic_intent.get("table_file", "3")
    explicit_query_type = detect_explicit_query_type(question, sample_ids, asv_ids)
    if explicit_query_type:
        query_type = explicit_query_type

    marker_reports = []
    for marker in markers:
        columns, source = load_table_columns(data_dir, marker, table_file)
        sample_status = [_resolve_sample_status(data_dir, sid, columns) for sid in sample_ids]
        resolved_pairs = [
            (st["requested_id"], st["actual_column"])
            for st in sample_status
            if st["present_as_column"]
        ]

        data = None
        if query_type == "asv_details" and asv_ids:
            details = [
                detail for asv_id in asv_ids
                if (detail := get_asv_detail(data_dir, marker, asv_id, table_file=table_file))
            ]
            if details:
                data = {"type": "asv_details", "rows": details}
        elif query_type == "species_matches" and species_name:
            matches = find_species(data_dir, marker, species_name, table_file=table_file)
            if matches:
                data = {
                    "type": "species_matches",
                    "species_name": species_name,
                    "rows": sorted(matches, key=lambda row: row["total_reads"], reverse=True)[:limit],
                }
        elif query_type == "sample_asvs" and resolved_pairs:
            groups = []
            for requested, resolved in resolved_pairs:
                rows = get_sample_asvs(data_dir, marker, resolved, limit, table_file=table_file)
                if rows:
                    groups.append({"requested": requested, "resolved": resolved, "rows": rows})
            if groups:
                data = {"type": "sample_asvs", "groups": groups}
        elif query_type == "sample_species" and resolved_pairs:
            groups = []
            for requested, resolved in resolved_pairs:
                rows = get_sample_species(data_dir, marker, resolved, limit, table_file=table_file)
                if rows:
                    groups.append({"requested": requested, "resolved": resolved, "rows": rows})
            if groups:
                data = {"type": "sample_species", "groups": groups}
        elif query_type == "sample_total_reads" and resolved_pairs:
            rows = []
            for requested, resolved in resolved_pairs:
                row = get_sample_total_reads(data_dir, marker, resolved, table_file=table_file)
                if row:
                    rows.append({"requested": requested, "resolved": resolved, **row})
            if rows:
                data = {"type": "sample_total_reads", "rows": rows}
        elif query_type == "top_asvs":
            rows = get_top_asvs(data_dir, marker, limit, table_file=table_file)
            if rows:
                data = {"type": "top_asvs", "rows": rows}
        elif query_type == "top_species":
            rows = get_top_species(data_dir, marker, limit, table_file=table_file)
            if rows:
                data = {"type": "top_species", "rows": rows}

        marker_reports.append(
            {
                "marker": marker,
                "table_file": table_file,
                "source": source,
                "available_sample_columns": columns,
                "sample_status": sample_status,
                "data": data,
            }
        )

    if not any(r["source"] for r in marker_reports) and not marker_reports:
        return None

    return {
        "intent": {
            "query_type": query_type,
            "table_file": table_file,
            "sample_ids": sample_ids,
            "asv_ids": asv_ids,
            "species_name": species_name,
            "markers": markers,
        },
        "markers": marker_reports,
    }


def format_table_entries(gathered) -> list[dict]:
    if not gathered:
        return []
    entries = []
    for report in gathered["markers"]:
        if report["source"] is None:
            entry_source = f"<{report['marker']}_table_file_{report['table_file']}_not_found>"
        else:
            entry_source = report["source"]
        lines = [
            f"Marker: {report['marker']}, table_file: `{report['table_file']}`",
            f"Sample columns present in this Excel file: {report['available_sample_columns']}",
        ]
        for status in report["sample_status"]:
            if status["present_as_column"]:
                lines.append(
                    f"Sample {status['requested_id']} present as column "
                    f"`{status['actual_column']}` in this file."
                )
            else:
                note = f"Sample {status['requested_id']} NOT present as a column in this file"
                if status["thesis_token"]:
                    note += f" (thesis token={status['thesis_token']})"
                if status["thesis_record"]:
                    note += f"; thesis record: {status['thesis_record']}"
                if status["evidence"]:
                    note += f"; resolution evidence: {status['evidence']}"
                lines.append(note + ".")
        data = report["data"]
        if data:
            t = data.get("type")
            lines.append(f"Fetched data ({t}):")
            if t == "sample_total_reads":
                for row in data["rows"]:
                    lines.append(
                        f"  - sample={row['requested']} (column={row['resolved']}): "
                        f"total_reads={row['total_reads']}"
                    )
            elif t == "asv_details":
                for row in data["rows"]:
                    lines.append(
                        f"  - {row['asv_id']}: species={row['species']}, "
                        f"total_reads={row['total_reads']}, identity={row['identity']}, "
                        f"target={row['target']}"
                    )
                    nonzero = {k: v for k, v in row["sample_reads"].items() if v}
                    if nonzero:
                        lines.append(f"    sample_reads: {nonzero}")
            elif t == "top_asvs":
                for row in data["rows"]:
                    lines.append(
                        f"  - {row['asv_id']}: species={row['species']}, "
                        f"total_reads={row['total_reads']}, identity={row['identity']}"
                    )
            elif t == "top_species":
                for row in data["rows"]:
                    lines.append(
                        f"  - {row['species']}: total_reads={row['total_reads']}, "
                        f"example_asv={row['example_asv']}"
                    )
            elif t == "species_matches":
                lines.append(f"  Searching: '{data['species_name']}'")
                for row in data["rows"]:
                    lines.append(
                        f"  - {row['asv_id']}: total_reads={row['total_reads']}, "
                        f"identity={row['identity']}, target={row['target']}"
                    )
            elif t == "sample_species":
                for grp in data["groups"]:
                    lines.append(f"  Sample {grp['requested']} (column={grp['resolved']}):")
                    for row in grp["rows"]:
                        lines.append(
                            f"    - {row['species']}: sample_reads={row['sample_reads']}, "
                            f"example_asv={row['example_asv']}"
                        )
            elif t == "sample_asvs":
                for grp in data["groups"]:
                    lines.append(f"  Sample {grp['requested']} (column={grp['resolved']}):")
                    for row in grp["rows"]:
                        lines.append(
                            f"    - {row['asv_id']}: species={row['species']}, "
                            f"sample_reads={row['sample_read_count']}, "
                            f"identity={row['identity']}"
                        )
        else:
            lines.append("No matching data fetched for this query type from this file.")

        entries.append({"kind": "table", "source": entry_source, "text": "\n".join(lines)})
    return entries


def detect_output_language(question: str) -> str:
    if re.search(r"[\u3040-\u30ff]", question):
        return "ja"
    if re.search(r"[\u4e00-\u9fff]", question):
        return "zh"
    return "en"


def get_output_labels(language: str) -> dict:
    if language == "ja":
        return {
            "details": "詳細",
            "estimated_species": "推定種",
            "sample_reads": "サンプル別 reads",
            "source": "出典",
            "top_asvs": "total reads が多い ASV",
            "top_species": "total reads が多い種",
            "sample_species": "サンプル内で検出された種",
            "sample_asvs": "サンプル内で reads が多い ASV",
            "sample_total_reads": "サンプルの total reads 合計",
            "assigned_asvs_prefix": "",
            "assigned_asvs_suffix": "と判断された ASV",
        }
    if language == "zh":
        return {
            "details": "详情",
            "estimated_species": "推定物种",
            "sample_reads": "各样本 reads",
            "source": "来源",
            "top_asvs": "total reads 最高的 ASV",
            "top_species": "total reads 最高的物种",
            "sample_species": "该样本中检出的物种",
            "sample_asvs": "该样本中 reads 最高的 ASV",
            "sample_total_reads": "该样本的 total reads 合计",
            "assigned_asvs_prefix": "判定为",
            "assigned_asvs_suffix": "的 ASV",
        }
    return {
        "details": "details",
        "estimated_species": "Estimated species",
        "sample_reads": "Sample reads",
        "source": "Source",
        "top_asvs": "Top ASVs by total reads",
        "top_species": "Top species by total reads",
        "sample_species": "Species detected in sample",
        "sample_asvs": "Top ASVs in sample",
        "sample_total_reads": "Total reads in sample",
        "assigned_asvs_prefix": "ASVs assigned to",
        "assigned_asvs_suffix": "",
    }


def format_table_answer(table_result: dict) -> str:
    question = table_result.get("question", "")
    language = detect_output_language(question)
    labels = get_output_labels(language)
    requirements = table_result.get("response_requirements") or {}
    output_mode = requirements.get("output_mode") or "prose"
    heading = {
        "ja": "構造化テーブル検索結果:",
        "zh": "结构化表格查询结果：",
    }.get(language, "Structured table query result:")
    lines = [] if output_mode == "markdown_table" else [heading]

    def append_markdown_header(headers: list[str]):
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join("---" for _ in headers) + "|")

    def display_sample_name(result_item: dict) -> str:
        requested = result_item.get("sample_id")
        resolved = result_item.get("resolved_sample_name") or requested
        if requested and resolved and requested != resolved:
            return f"{requested} ({resolved})"
        return requested or resolved or ""

    for result in table_result["results"]:
        result_type = result["type"]

        if result_type == "asv_details":
            if output_mode == "markdown_table":
                headers = ["ASV", labels["estimated_species"], "Identity", "Total reads"]
                if requirements.get("include_target"):
                    headers.append("BLAST target")
                append_markdown_header(headers)
                for row in result["data"]:
                    values = [row["asv_id"], row["species"], str(row["identity"]), str(row["total_reads"])]
                    if requirements.get("include_target"):
                        values.append(str(row["target"]))
                    lines.append("| " + " | ".join(values) + " |")
                continue
            for row in result["data"]:
                lines.append(f"\n[{row['marker']}] {row['asv_id']} {labels['details']}")
                lines.append(f"- {labels['estimated_species']}: {row['species']}")
                lines.append(f"- BLAST target: {row['target']}")
                lines.append(f"- Identity: {row['identity']}")
                lines.append(f"- Total reads: {row['total_reads']}")
                nonzero = {
                    sample: reads
                    for sample, reads in row["sample_reads"].items()
                    if reads
                }
                if nonzero:
                    lines.append(f"- {labels['sample_reads']}:")
                    for sample, reads in sorted(nonzero.items(), key=lambda item: item[1], reverse=True):
                        lines.append(f"  - {sample}: {reads}")
                lines.append(f"- {labels['source']}: {row['source']}")

        if result_type == "top_asvs":
            if output_mode == "markdown_table":
                append_markdown_header(["ASV", labels["estimated_species"], "Total reads", "Identity"])
                for row in result["data"]:
                    lines.append(
                        f"| {row['asv_id']} | {row['species']} | {row['total_reads']} | {row['identity']} |"
                    )
                continue
            lines.append(f"\n[{result['marker']}] {labels['top_asvs']}")
            for row in result["data"]:
                lines.append(
                    f"- {row['asv_id']}: {row['species']}; "
                    f"total_reads={row['total_reads']}; identity={row['identity']}"
                )

        if result_type == "top_species":
            if output_mode == "markdown_table":
                append_markdown_header([labels["estimated_species"], "Total reads", "Example ASV"])
                for row in result["data"]:
                    lines.append(f"| {row['species']} | {row['total_reads']} | {row['example_asv']} |")
                continue
            lines.append(f"\n[{result['marker']}] {labels['top_species']}")
            for row in result["data"]:
                lines.append(
                    f"- {row['species']}: total_reads={row['total_reads']}; "
                    f"example_asv={row['example_asv']}"
                )

        if result_type == "species_matches":
            if output_mode == "markdown_table":
                append_markdown_header(["ASV", "Total reads", "Identity", "BLAST target"])
                for row in result["data"]:
                    lines.append(
                        f"| {row['asv_id']} | {row['total_reads']} | {row['identity']} | {row['target']} |"
                    )
                continue
            if language == "ja":
                lines.append(f"\n[{result['marker']}] {result['species_name']} {labels['assigned_asvs_suffix']}")
            elif language == "zh":
                lines.append(
                    f"\n[{result['marker']}] {labels['assigned_asvs_prefix']} "
                    f"{result['species_name']} {labels['assigned_asvs_suffix']}"
                )
            else:
                lines.append(f"\n[{result['marker']}] {labels['assigned_asvs_prefix']} {result['species_name']}")
            for row in result["data"]:
                lines.append(
                    f"- {row['asv_id']}: total_reads={row['total_reads']}; "
                    f"identity={row['identity']}; target={row['target']}"
                )

        if result_type == "sample_species":
            if output_mode == "markdown_table":
                append_markdown_header([labels["estimated_species"], "Sample reads", "Example ASV"])
                for row in result["data"]:
                    lines.append(f"| {row['species']} | {row['sample_reads']} | {row['example_asv']} |")
                continue
            lines.append(f"\n[{result['marker']}] {display_sample_name(result)} {labels['sample_species']}")
            for row in result["data"]:
                lines.append(
                    f"- {row['species']}: sample_reads={row['sample_reads']}; "
                    f"example_asv={row['example_asv']}"
                )

        if result_type == "sample_asvs":
            if output_mode == "markdown_table":
                append_markdown_header(["ASV", labels["estimated_species"], "Sample reads", "Identity"])
                for row in result["data"]:
                    lines.append(
                        f"| {row['asv_id']} | {row['species']} | {row['sample_read_count']} | {row['identity']} |"
                    )
                continue
            lines.append(f"\n[{result['marker']}] {display_sample_name(result)} {labels['sample_asvs']}")
            for row in result["data"]:
                lines.append(
                    f"- {row['asv_id']}: {row['species']}; "
                    f"sample_reads={row['sample_read_count']}; identity={row['identity']}"
                )

        if result_type == "sample_total_reads":
            row = result["data"]
            if output_mode == "markdown_table":
                append_markdown_header(["Sample", "Total reads", labels["source"]])
                lines.append(f"| {display_sample_name(result)} | {row['total_reads']} | {row['source']} |")
                continue
            lines.append(f"\n[{row['marker']}] {display_sample_name(result)} {labels['sample_total_reads']}")
            lines.append(f"- total_reads={row['total_reads']}")
            lines.append(f"- {labels['source']}: {row['source']}")

        if result_type == "sequence_mapping_missing":
            sample_text = []
            for resolution in result["resolutions"]:
                detail = resolution.requested_id
                if resolution.thesis_token:
                    detail += f" [token={resolution.thesis_token}]"
                metadata = summarize_thesis_record(resolution.thesis_record)
                if metadata:
                    detail += f" ({metadata})"
                if resolution.evidence:
                    detail += f"; evidence={', '.join(resolution.evidence)}"
                sample_text.append(detail)
            joined = "<br>".join(sample_text) if output_mode == "markdown_table" else ", ".join(sample_text)
            if output_mode == "markdown_table":
                append_markdown_header(["Status", "Samples", labels["source"]])
                lines.append(f"| mapping unavailable | {joined} | {result['source']} |")
                continue
            lines.append(f"\n[{result['marker']}] mapping unavailable")
            lines.append(f"- samples: {joined}")
            lines.append(f"- {labels['source']}: {result['source']}")

        if result_type == "sample_not_found":
            sample_text = ", ".join(result["sample_ids"])
            if output_mode == "markdown_table":
                append_markdown_header(["Status", "Samples", labels["source"]])
                lines.append(f"| sample not found | {sample_text} | {result['source']} |")
                continue
            lines.append(f"\n[{result['marker']}] sample not found")
            lines.append(f"- samples: {sample_text}")
            lines.append(f"- {labels['source']}: {result['source']}")

    return "\n".join(lines)
